# IMPORTANDO MÓDULOS E PACOTES
import logging
import openpyxl
import shutil
import os
import traceback
import pandas as pd
import logging
from transform import Transform


class Load_camp_rank_ativ:
    def __init__(self):
        # Carrega os dataframes necessários ao instanciar a classe
        transform_instance = Transform()
        self.df_final_ativacoes, self.df_final_cancelamentos_integrais, self.df_ativacoes_dia_anterior_ranking = transform_instance.transforming_files()

    def get_last_row(self, sheet):
        for row in range(sheet.max_row, 0, -1):
            if sheet.cell(row, 1).value is not None:
                return row + 1
        return 1

    def clear_sheet(self, sheet):
        max_row = sheet.max_row
        if max_row > 1:
            sheet.delete_rows(2, max_row)

    def load_files(self):
        try:
            # Usa os dataframes carregados na instância
            df_final_ativacoes = self.df_final_ativacoes
            df_cancelamentos = self.df_final_cancelamentos_integrais

            template = r"C:\Users\raphael.almeida\Documents\Processos\placas_acompanhamento\template\placas_movimentacoes.xlsx"
            wb = openpyxl.load_workbook(filename=template)
            ws1 = wb['ATIVAÇÕES']
            ws2 = wb['CANCELAMENTOS']

            # REFORMATANDO DATAFRAMES APENAS COM OS REGISTROS QUE NÃO ESTÃO NA PLANILHA (EVITANDO INSERÇÃO DE DUPLICATAS)
            df_validacao_canc = pd.DataFrame(ws2.values)

            if not df_cancelamentos.empty:
                for index, row in df_cancelamentos.iterrows():
                    if row[0] in df_validacao_canc.values:
                        df_cancelamentos.drop(index, inplace=True)

            # ENCONTRANDO A ÚLTIMA LINHA VAZIA NA PLANILHA PARA A INSERÇÃO DOS DADOS
            last_row_canc = self.get_last_row(ws2)

            # LIMPANDO OS DADOS NA ABA 'ATIVAÇÕES'
            self.clear_sheet(ws1)

            # ADICIONANDO OS DADOS NA ABA 'ATIVAÇÕES'
            if not df_final_ativacoes.empty:
                for r_idx, row in enumerate(df_final_ativacoes.values, start=2):
                    for c_idx, value in enumerate(row, start=1):
                        if pd.isna(value):
                            value = None
                        ws1.cell(row=r_idx, column=c_idx, value=value)
            else:
                logging.info('\n ----------------------------------------------------------------------------------')
                logging.info('\n Não há registros de ativações para serem adicionados ao ranking')

            # ADICIONANDO OS DADOS NA ABA 'CANCELAMENTOS'
            if not df_cancelamentos.empty:
                for r_idx, row in enumerate(df_cancelamentos.values, start=last_row_canc):
                    for c_idx, value in enumerate(row, start=1):
                        if pd.isna(value):
                            value = None
                        ws2.cell(row=r_idx, column=c_idx, value=value)
            else:
                logging.info('\n ----------------------------------------------------------------------------------')
                logging.info('\n Não há registros de cancelamentos para serem adicionados na planilha. Registros já existentes na planilha')

            # SALVANDO PLANILHA
            wb.save(r"C:\Users\raphael.almeida\Documents\Processos\placas_acompanhamento\template\placas_movimentacoes.xlsx")
            wb.close()

            file_path = r"C:\Users\raphael.almeida\Documents\Processos\placas_acompanhamento\template\placas_movimentacoes.xlsx"
            destination_dir = r"C:\Users\raphael.almeida\OneDrive - Grupo Unus\analise de dados - Arquivos em excel"
            destination_path = os.path.join(destination_dir, os.path.basename(file_path))

            shutil.copy(file_path, destination_path)

            logging.info('\n ----------------------------------------------------------------------------------')
            logging.info('\n Processo de Carregamento de Dados concluido com sucesso!')

        except Exception as e:
            logging.info('\n ----------------------------------------------------------------------------------')
            logging.info(f'\n Falha ao Carregar Relatorio Final: {e}')
            logging.info(traceback.format_exc())


if __name__ == '__main__':
    load = Load_camp_rank_ativ()
    load.load_files()





    # extract_instance = Extract() #Fazer isso quando a class contiver instâncias 

    # df_ativacoes = extract_instance.extract_all_ativacoes() #salvando em variáveis os métodos da instância de classe
    # df_cancelamentos = extract_instance.extract_all_cancelamentos() 
    # df_conferencia = extract_instance.extract_conf_boards()
   
    # Criando uma instância da classe Transform, pois não é uma classe estática (usa self)
    # transform_instance = Transform()  # Passando um dicionário vazio como config

    # Chamando o método transforming_files da classe Transform, que retorna os DataFrames processados    
    # df_final_ativacoes, df_cancelamentos = transform_instance.transforming_files(df_ativacoes, df_cancelamentos, df_conferencia)

    # Criando uma instância da classe Load_camp_rank_ativ, pois ela não é estática (usa self)
    # load_instance = Load_camp_rank_ativ()

    # Chamando o método load_files da instância da classe Load_camp_rank_ativ, passando os DataFrames processados
    # load_instance.load_files(df_final_ativacoes, df_cancelamentos)







    
